import openpyxl
from datetime import date, timedelta

# Carregar a planilha
workbook = openpyxl.load_workbook('Plano_Sobre_Objetivos_e_Metas.xlsm')
sheet_name = "Bd_Dados"  # Use a planilha ativa ou substitua pelo nome da planilha
sheet = workbook[sheet_name]

# Loop pelas linhas da planilha (começando da segunda linha, pois a primeira é cabeçalho)
for row in sheet.iter_rows(min_row=2, values_only=True):
    objetivo_meta = row[0]
    politicas = row[1:9]  # Coluna B a I
    descricao_acao = row[9]
    data_prevista = row[10]
    data_realizado = row[12]
    status = row[13]
    avaliacao_eficacia = row[14]
    observacao = row[15]
    data_obs = row[16]

    # Fazer análises ou processamentos com os dados
    if status == 'Em andamento' and data_prevista < date.today():
        status = 'Em atraso'
        sheet.cell(row=row[0].row, column=13, value=status)  # Atualizar o status na planilha

    if status == 'Realizado' and avaliacao_eficacia != 'Eficaz':
        observacao = 'Ação realizada, mas não eficaz'
        sheet.cell(row=row[0].row, column=15, value=observacao)  # Atualizar observação

# Salvar a planilha com as atualizações
workbook.save('sua_planilha_atualizada.xlsx')
